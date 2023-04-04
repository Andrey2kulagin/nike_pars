import openpyxl
import glob


def write_excel_row(data_row, last_row, worksheet):
    for col_idx, cell_value in enumerate(data_row, start=1):
        worksheet.cell(row=last_row + 1, column=col_idx, value=cell_value)


def write_to_file(data_row, filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    last_row = worksheet.max_row
    write_excel_row(data_row, last_row, worksheet)


files = glob.glob('day_1/1cards_data*.xlsx')
for file in files:
    print(file)
    from_workbook = openpyxl.load_workbook(filename=file)
    # выбор нужного листа
    from_sheet = from_workbook.active
    to_workbook = openpyxl.load_workbook(filename='day_1/data.xlsx')
    to_sheet = to_workbook.active
    rows = from_sheet.iter_rows(values_only=True)
    is_first = True
    for row in rows:
        if is_first:
            is_first = False
            continue
        last_row = to_sheet.max_row
        write_excel_row(row, last_row, to_sheet)
    to_workbook.save('day_1/data.xlsx')
