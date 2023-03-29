import openpyxl


def get_product_row(card_data, last_row):
    data_row = []
    data_row.append(card_data['model_sub_name'])
    data_row.append("Товар")
    product_number = str(last_row + 1)
    data_row.append(product_number)
    data_row.append(card_data["model_main_name"])
    link_string = link_array_to_string(card_data["photos"])
    data_row.append(link_string)
    data_row.append(card_data["article"])
    data_row.append(card_data["price"])
    data_row.append("доллар")
    data_row.append("")
    data_row.append("")
    return data_row


def data_for_excel(card_data, last_row):
    product_row = get_product_row()


def link_array_to_string(link_array):
    link_str = ""
    for link in link_array:
        link_str += (link + ' ; ')
    return link_str


def write_excel_row(data_row, last_row, worksheet):
    for col_idx, cell_value in enumerate(data_row, start=1):
        worksheet.cell(row=last_row + 1, column=col_idx, value=cell_value)


card_data = {'article': 'DV9956-118', 'photos': [
    'https://static.nike.com/a/images/t_PDP_1280_v1/f_auto,q_auto:eco/18e48a78-6740-4f18-a103-04c299e37be5/air-jordan-2-retro-low-mens-shoes-LVlsNr.png',
    'https://static.nike.com/a/images/t_PDP_1280_v1/f_auto,q_auto:eco/7176fed1-c14c-4e2f-8ad7-6031170080d3/air-jordan-2-retro-low-mens-shoes-LVlsNr.png',
    'https://static.nike.com/a/images/t_PDP_1280_v1/f_auto,q_auto:eco/0f2aaaba-2361-4d32-9432-2b085403116d/air-jordan-2-retro-low-mens-shoes-LVlsNr.png',
    'https://static.nike.com/a/images/t_PDP_1280_v1/f_auto,q_auto:eco/a57a6814-19f6-49bb-8497-1a25aecb6db0/air-jordan-2-retro-low-mens-shoes-LVlsNr.png',
    'https://static.nike.com/a/images/t_PDP_1280_v1/f_auto,q_auto:eco/7f11e4a2-dd9a-4f2d-8af0-23d77d7c5059/air-jordan-2-retro-low-mens-shoes-LVlsNr.png',
    'https://static.nike.com/a/images/t_PDP_1280_v1/f_auto,q_auto:eco/270e8f41-b6bb-46bd-a59d-9356d037d821/air-jordan-2-retro-low-mens-shoes-LVlsNr.png',
    'https://static.nike.com/a/images/t_PDP_1280_v1/f_auto,q_auto:eco/de28eb1c-ac7d-4877-aa74-b21e0921a5bd/air-jordan-2-retro-low-mens-shoes-LVlsNr.png',
    'https://static.nike.com/a/images/t_PDP_1280_v1/f_auto,q_auto:eco/b987fe64-8019-4b64-80dc-dc26e7f454dc/air-jordan-2-retro-low-mens-shoes-LVlsNr.png'],
             'model_main_name': 'Air Jordan 2 Retro Low', 'model_sub_name': "Men's Shoes", 'price': '$150',
             'size': ['M 3.5 / W 5', 'M 4 / W 5.5', 'M 4.5 / W 6', 'M 5 / W 6.5', 'M 5.5 / W 7', 'M 6 / W 7.5',
                      'M 6.5 / W 8', 'M 7 / W 8.5', 'M 7.5 / W 9', 'M 8 / W 9.5', 'M 8.5 / W 10', 'M 9 / W 10.5',
                      'M 9.5 / W 11', 'M 10 / W 11.5', 'M 10.5 / W 12', 'M 11 / W 12.5', 'M 11.5 / W 13',
                      'M 12 / W 13.5',
                      'M 12.5 / W 14', 'M 13 / W 14.5', 'M 14 / W 15.5', 'M 15 / W 16.5', 'M 16 / W 17.5',
                      'M 17 / W 18.5',
                      'M 18 / W 19.5']}

workbook = openpyxl.load_workbook('cards_data.xlsx')
worksheet = workbook.active
last_row = worksheet.max_row

data_row = get_product_row(card_data, last_row)
write_excel_row(data_row, last_row, worksheet)
last_row += 1
if len(card_data["count"]):
    #обработка всех размеров и их запись
    pass

# Сохраняем файл
workbook.save('cards_data.xlsx')
