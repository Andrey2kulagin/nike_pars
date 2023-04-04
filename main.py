import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
from datetime import datetime
import os


def counter():
    count = 1374  # значение счётчика

    def inner():
        nonlocal count  # указываем, что переменная count находится во внешней области видимости
        count += 1  # увеличиваем значение счётчика на 1
        return count  # возвращаем текущее значение счётчика

    return inner


my_counter = counter()


def get_product_row(card_data, product_number):
    data_row = []
    data_row.append(card_data['model_sub_name'])
    data_row.append("Товар")
    data_row.append(str(product_number))
    data_row.append(card_data["model_main_name"])
    photo_string = link_array_to_string(card_data["photos"])
    data_row.append(photo_string)
    if len(card_data['size']) == 0:
        data_row.append(card_data['article'])
    else:
        data_row.append("")
    data_row.append(card_data["price"])
    data_row.append("доллар")
    data_row.append("")
    data_row.append("")
    return data_row


def link_array_to_string(link_array):
    link_str = ""
    for link in link_array:
        link_str += (link + ' ; ')
    return link_str


def write_excel_row(data_row, last_row, worksheet):
    for col_idx, cell_value in enumerate(data_row, start=1):
        worksheet.cell(row=last_row + 1, column=col_idx, value=cell_value)


def get_size_color_row(size, main_product_num, article, photos):
    data_row = [""]
    data_row.append("Модификация")
    data_row.append(str(my_counter()))
    data_row.append("")
    photo_string = link_array_to_string(photos)
    data_row.append(photo_string)
    data_row.append(article)
    data_row.append("")
    data_row.append("")
    data_row.append(str(main_product_num))
    data_row.append(size)
    return data_row


def is_sizes_empty(cards_data):
    card_len = 0
    for card_data in cards_data:
        card_len += len(cards_data[card_data]["size"])
    return card_len == 0


def get_product_row_without_sizes(card_data):
    data_row = []
    data_row.append(card_data['model_sub_name'])
    data_row.append("Товар")
    product_number = str(my_counter())
    data_row.append(product_number)
    data_row.append(card_data["model_main_name"])
    photo_string = link_array_to_string(card_data["photos"])
    data_row.append(photo_string)
    data_row.append("")
    data_row.append("")
    data_row.append("")
    data_row.append("")
    data_row.append("")
    return data_row


def get_article_row(main_product_num, card_data):
    data_row = [""]
    data_row.append("Модификация")
    data_row.append(str(my_counter()))
    data_row.append("")
    photo_string = link_array_to_string(card_data["photos"])
    data_row.append(photo_string)
    data_row.append(card_data["article"])
    data_row.append(card_data["price"])
    data_row.append("доллар")
    data_row.append(str(main_product_num))
    data_row.append("")
    return data_row


def write_to_file(cards_data, filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    last_row = worksheet.max_row
    main_product_num = None
    is_product_add = False
    if not is_sizes_empty(cards_data):
        for card_data in cards_data:
            if not is_product_add:
                main_product_num = str(my_counter())
                data_row = get_product_row(cards_data[card_data], main_product_num)
                write_excel_row(data_row, last_row, worksheet)
                last_row += 1
                is_product_add = True
            if len(cards_data[card_data]["size"]):
                # обработка всех размеров и их запись
                for size in cards_data[card_data]["size"]:
                    data_row = get_size_color_row(size, main_product_num, cards_data[card_data]["article"],
                                                  cards_data[card_data]["photos"])
                    write_excel_row(data_row, last_row, worksheet)
                    print("записана строка:", last_row)
                    last_row += 1
            # Сохраняем файл
    else:
        for card_data in cards_data:
            if not is_product_add:
                data_row = get_product_row_without_sizes(cards_data[card_data])
                write_excel_row(data_row, last_row, worksheet)
                main_product_num = last_row
                last_row += 1
                is_product_add = True
            # добавляем артикул с фотками и т.д.
            data_row = get_article_row(main_product_num, cards_data[card_data])
            write_excel_row(data_row, last_row, worksheet)
            print("записана строка:", last_row)
            last_row += 1

    workbook.save(filename)
    return last_row


def get_card_color_info(driver, data, is_sold_out):
    while True:
        try:
            images_block = driver.find_element(By.ID, "experience-wrapper")
            break
        except Exception as e:
            if "SBOX_FATAL_MEMORY_EXCEEDED" in str(e):
                driver.refresh()
    images_block1 = images_block.find_element(By.CLASS_NAME, "css-1rayx7p")
    buttons = images_block1.find_elements(By.CLASS_NAME, "css-du206p")

    photos = []
    for button in buttons:
        picture = button.find_elements(By.TAG_NAME, "picture")[1]
        img = picture.find_element(By.TAG_NAME, "img")
        href = img.get_attribute('src')
        photos.append(href)
        # print(href)
    data["photos"] = photos
    model_main_name = driver.find_element(By.XPATH, "//*[@id='pdp_product_title']")
    data["model_main_name"] = model_main_name.text
    # print(model_main_name.text)
    model_sub_name_block = driver.find_element(By.CSS_SELECTOR, '.d-lg-ib.mb0-sm.u-full-width.css-3rkuu4.css-1mzzuk6')
    model_sub_name = driver.find_element(By.CSS_SELECTOR, '.headline-5.pb1-sm.d-sm-ib').text
    data["model_sub_name"] = model_sub_name
    # print(model_sub_name)
    price = model_sub_name_block.find_element(By.CSS_SELECTOR,
                                              ".product-price").text[1:]
    data["price"] = price
    data["size"] = []
    if not is_sold_out and not is_coming_soon(driver):
        try:
            sizes_block = driver.find_element(By.CSS_SELECTOR, ".mt5-sm.mb3-sm.body-2")
            active_sizes_inputs = sizes_block.find_elements(By.CSS_SELECTOR, "input:not([disabled])")
            sizes = []
            for input in active_sizes_inputs:
                input_id = input.get_attribute('id')
                size = sizes_block.find_element(By.CSS_SELECTOR, f"label[for='{input_id}']").text
                sizes.append(size)
                # (size)
            if not len(sizes):
                raise Exception("Размеры не спарсились")
            data["size"] = sizes
        except Exception as e:
            f = open("bags.txt", 'a')
            f.write("\n\nНачало бага\n")
            f.write(f"Размеры не спарсились почему-то, проверь артикул   {data['article']}")
            f.write(f"exception {e}")
            f.write(f"\n\n")
            f.close()


def is_coming_soon(driver):
    coming_soon = driver.find_elements(By.CLASS_NAME, 'sold-out')
    is_coming_soon_eq = len(coming_soon)
    coming_soon = driver.find_elements(By.CSS_SELECTOR,
                                       '.bg-primary-grey.ta-sm-c.mr1-md.ml0-md.mr4-sm.ml4-sm.pt4-sm.pb4-sm.mt5-md.mt8-sm.mb4-sm.mb4-md')
    is_coming_soon_eq += len(coming_soon)
    return is_coming_soon_eq


def cookie_allow(driver):
    card_link = "https://www.nike.com/t/air-jordan-1-mid-se-mens-shoes-Zn07hL/DV1308-104"
    driver.get(card_link)
    while True:
        try:
            time.sleep(3)
            element = driver.find_element(By.CLASS_NAME, "nds-dialog__actions")
            element.click()
            break
        except:
            driver.refresh()


def get_card_info(driver, card_link, card_data):
    data = {}
    driver.get(card_link)
    colors = driver.find_elements(By.CSS_SELECTOR, ".css-b8rwz8.tooltip-component-container")
    articles = []
    for color in colors:
        color_input = color.find_element(By.NAME, "pdp-colorpicker")
        color_article = color_input.get_attribute('data-style-color')
        is_sold_out = color_input.get_attribute('data-nr-sold-out')
        articles.append({"article": color_article, "is_sold_out": is_sold_out == 'true'})
    right_slash_index = card_link.rfind('/')
    base_link = card_link[:(right_slash_index + 1)]
    if len(colors) == 0:
        alone_article = card_link[(right_slash_index + 1):].strip('\n')
        articles.append({"article": alone_article, "is_sold_out": False})
    for article in articles:
        driver.get(base_link + article["article"])
        data["article"] = article["article"]
        # print(article)
        get_card_color_info(driver, data, article["is_sold_out"])
        card_data[article['article']] = dict(data)


def gen_filename(mask, files_count):
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S").replace(':', "")
    return mask + '_' + str(files_count) + '_' + current_time + '.xlsx'


def create_new_excel(filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    data_row = ["Группы", "Тип", "Код", "Наименование", "Изображение", "Характеристика:Артикул", "Цена: Цена продажи",
                "Валюта (Цена продажи)", "Код товара модификации", "Характеристика:Размер"]
    write_excel_row(data_row, 0, sheet)
    workbook.save(filename=filename)


def rename_old_datafile(old_name):
    new_name = '1' + old_name
    os.rename(old_name, new_name)


# Словарь с данными по карточке
data = {}
chrome_options = Options()
# неробот
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option("useAutomationExtension", False)
# chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)
cookie_allow(driver)
excel_filename_mask = "cards_data"
files_count = 4
filename = gen_filename(excel_filename_mask, files_count)
create_new_excel(filename)
max_str_count = 300
with open("links1_with_duplicate.txt", 'r', encoding='utf-8') as card_links:
    for card_link in card_links:
        print(card_link)
        card_data = {}
        try:
            get_card_info(driver, card_link, card_data)
        except Exception as e:
            f = open("bags.txt", 'a')
            f.write("\n\nНачало бага\n")
            f.write(f"link:   {card_link}")
            f.write(f"exception {e}")
            f.write(f"\n\n")
            f.close()
        cure_str_count = write_to_file(card_data, filename)
        if cure_str_count > max_str_count:
            files_count += 1
            rename_old_datafile(filename)
            filename = gen_filename(excel_filename_mask, files_count)
            create_new_excel(filename)

driver.quit()
